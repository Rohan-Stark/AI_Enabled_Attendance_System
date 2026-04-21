"""
Export Service — generates Excel attendance reports using openpyxl.
"""
import os
import logging
from typing import Optional
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import current_app
from app.services.attendance_service import get_subject_attendance_summary
from app.models.subject import Subject

logger = logging.getLogger(__name__)

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def export_attendance_to_excel(subject_id: int) -> Optional[str]:
    """
    Generate an Excel report for a subject's attendance.

    Returns:
        File path of the generated Excel file, or None on error.
    """
    subject = Subject.query.get(subject_id)
    if not subject:
        return None

    summary = get_subject_attendance_summary(subject_id)
    threshold = current_app.config.get("LOW_ATTENDANCE_THRESHOLD", 75)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ── Title row ──────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance Report — {subject.name} ({subject.code})"
    title_cell.font = Font(name="Calibri", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Headers ────────────────────────────────────────────
    headers = ["S.No", "Enrollment No", "Student Name", "Present", "Total Sessions", "Attendance %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # ── Data rows ──────────────────────────────────────────
    for i, record in enumerate(summary, 1):
        row = i + 4
        ws.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws.cell(row=row, column=2, value=record["enrollment_no"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=record["name"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=record["present"]).border = THIN_BORDER
        ws.cell(row=row, column=5, value=record["total"]).border = THIN_BORDER

        pct_cell = ws.cell(row=row, column=6, value=f"{record['percentage']}%")
        pct_cell.border = THIN_BORDER
        pct_cell.alignment = Alignment(horizontal="center")

        # Highlight low attendance
        if record["percentage"] < threshold:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = LOW_FILL

    # ── Auto-width columns ────────────────────────────────
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # ── Save ───────────────────────────────────────────────
    export_dir = current_app.config.get("EXPORT_DIR", "exports")
    os.makedirs(export_dir, exist_ok=True)
    filename = f"attendance_{subject.code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    logger.info(f"Excel report exported: {filepath}")
    return filepath
